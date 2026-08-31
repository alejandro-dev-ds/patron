def ejecutar_patron(
    historico_path,
    proyectos_path,
    correcciones_path,
    plantilla
):

    df_hist_horas = pd.read_excel(
        historico_path,
        engine="openpyxl"
    )

    df_projects = pd.read_excel(
        proyectos_path,
        engine="openpyxl"
    )

    df_projects = df_projects.rename(columns={"ACRÓNIMO": "PROYECTO"})
    df_projects = df_projects[df_projects["ESTADO"]=="CERRADO"]

    
    lista_planificaciones = [
    "CORTE Y ENSAMBLAJE DE CARRILES Y BANDEJAS",
    "NIVELACIÓN",
    "INSTALACIÓN CARRILES DE SOPORTACIÓN",
    "INSTALACIÓN DE BANDEJAS",
    "MECANIZADO DE ELEMENTOS DE INSTALACIÓN MECÁNICA",
    "INTRODUCCIÓN/FIJACIÓN de EQUIPOS",
    "ENSAMBLAJE DE CUADROS CUADRISTA",
    "ENSAMBLAJE DE EQUIPOS (CUADRO PRINCIPAL Y UPS)",
    "CADENAS PORTACABLES",
    "INSTALACIÓN DE SUELO",
    "PANELADO",
    "INSTALACIÓN DE PUERTAS",
    "INSTALACIÓN ROXTEC",
    "INSTALACIÓN DE ELEMENTOS y/o EQUIPOS",
    "CERRAMIENTO DE ALUMINIO",
    "PERFILERíA EXTERIOR",
    "CORTE Y/O PREPARACIÓN DE CABLE",
    "MECANIZADO DE ELEMENTOS DE INSTALACIÓN ELÉCTRICA",
    "SISTEMA DE TIERRAS",
    "SISTEMA DE ILUMINACIÓN/EMERGENCIAS",
    "INSTALACIÓN DE CABLE",
    "CONEXIONADO DE EQUIPOS",
    "ETIQUETADO DE CABLE Y EQUIPOS",
    "TORQUEO",
    "TRANSFORMADOR",
    "TRABAJOS DE MEDIA TENSIÓN",
    "TRABAJOS DE BT EN MT",
    "PRUEBAS DE MEDIA TENSION",
    "MONTAJE DE EQUIPOS DE MONITORIZACIÓN",
    "INSTALACIÓN DE SEÑALES",
    "INSTALACIÓN DE UTP",
    "CONTROL DE ACCESO",
    "PRUEBAS ELECTRICAS SIN TENSIÓN",
    "PRUEBAS ELECTRICAS CON TENSIÓN",
    "SOPORTE A COMMISSIONING",
    "TRABAJOS DE INSTALACIÓN PCI",
    "REALIZACIÓN DE PRUEBAS PCI",
    "MONTAJE DE TUBERÍAS",
    "TRABAJOS DE CONEXIONADO Y VALVULERIA",
    "TRABAJOS DE FORRADO DE TUBERÍAS",
    "REALIZACIÓN DE PRUEBAS COOLING",
    "MODIFICACIONES FAT/FOK/WITNESS TEST",
    "DESENSAMBLAJE INSTALACIÓN ELÉCTRICA",
    "DESENSAMBLAJE INSTALACIÓN MECÁNICA",
    "LIMPIEZA FINAL CUADROS",
    "CIERRE DE CUADROS Y PENDIENTES ELÉCTRICOS",
    "CIERRE DE CUADROS Y PENDIENTES MECÁNICOS",
    "LIMPIEZA FINAL",
    "REPASOS DE PINTURA",
    "EMBALAJE"
    ]

    def leer_correcciones_visual(
    fichero,
    lista_planificaciones
):
        df_raw = pd.read_excel(
        fichero,
        header=None
        )

        registros = []

        bloque_actual = None

        estadisticas = [
            "MEDIANA",
            "PROMEDIO",
            "MAX",
            "MIN"
        ]

        for _, fila in df_raw.iterrows():

            valor = fila.iloc[0]

            if pd.isna(valor):
                continue

            texto = str(valor).strip()

            # fila bloque
            if (
                texto not in estadisticas
                and texto != "PROYECTO"
                and pd.isna(fila.iloc[1])
            ):
                bloque_actual = texto
                continue

            # cabecera
            if texto == "PROYECTO":
                continue

            # estadísticas
            if texto in estadisticas:
                continue

            proyecto = texto

            horas = fila.iloc[
                1:1+len(lista_planificaciones)
            ].tolist()

            registros.append(
                [bloque_actual, proyecto] + horas
            )

        columnas = (
            ["FAMILIA_TAMAÑO", "PROYECTO"]
            + lista_planificaciones
        )

        return pd.DataFrame(
            registros,
            columns=columnas
        )


    df_correcciones = leer_correcciones_visual(
        correcciones_path,
        lista_planificaciones
    )

    df_corr_largo = (
        df_correcciones
        .melt(
            id_vars="PROYECTO",
            var_name="PLANIFICACIÓN",
            value_name="HORAS DEDICADAS"
        )
    )

    df_corr_plano = df_corr_largo[
        ~df_corr_largo["PROYECTO"].isin(
            ["MEDIANA", "PROMEDIO", "MAX", "MIN"]
        )
    ]

    claves = ["PROYECTO", "PLANIFICACIÓN"]

    df_hist_horas = (
        df_hist_horas
        .merge(
            df_corr_largo[claves].drop_duplicates(),
            on=claves,
            how="left",
            indicator=True
        )
    )

    df_hist_horas = (
        df_hist_horas[
            df_hist_horas["_merge"] == "left_only"
        ]
        .drop(columns="_merge")
    )

    df_hist_horas = pd.concat(
        [df_hist_horas, df_corr_largo],
        ignore_index=True
    )

    df_fin = df_hist_horas.merge(df_projects,on="PROYECTO")

    familias = ["AIO","DH","PM","MO","SKID","BM"]

    df_fin = df_fin[df_fin["FAMILIA"].isin(familias)]

    df_fin["FAMILIA_TAMAÑO"] = (
        df_fin["FAMILIA"].astype(str) + "_" + df_fin["TAMAÑO"].astype(str)
    )

    df_resumen = (
        df_fin.groupby(
            ["FAMILIA_TAMAÑO", "PROYECTO", "PLANIFICACIÓN"],
            as_index=False
        )["HORAS DEDICADAS"]
        .sum()
    )

    tabla = pd.pivot_table(
        df_resumen,
        values="HORAS DEDICADAS",
        index=["FAMILIA_TAMAÑO", "PROYECTO"],
        columns="PLANIFICACIÓN",
        aggfunc="sum",
        fill_value=0
    )

    

    for col in lista_planificaciones:
        if col not in tabla.columns:
            tabla[col] = 0

    tabla_fil = tabla.reindex(columns=lista_planificaciones)

    # ==========================================================
    # GENERAR DICCIONARIO PARA EL EXCEL
    # ==========================================================

    datos = {}

    grupos = (
        tabla_fil.index.droplevel("PROYECTO")
        .unique()
    )

    for familia_tamaño in grupos:

        bloque = tabla_fil.loc[(familia_tamaño)]

        nombre_bloque = f"{familia_tamaño}"

        proyectos = []

        for proyecto, fila in bloque.iterrows():

            proyectos.append(
                [proyecto] + fila.tolist()
            )

        datos[nombre_bloque] = proyectos


    PLANTILLA = "ejemplo_tabla_patron.xlsx"
    SALIDA = "trabajo_patron_horas.xlsx"

    # =====================================================
    # FUNCIONES
    # =====================================================

    def copiar_estilo_fila(ws, fila_origen, fila_destino, max_col):
        """
        Copia formato completo de una fila a otra.
        """

        for col in range(1, max_col + 1):

            origen = ws.cell(fila_origen, col)
            destino = ws.cell(fila_destino, col)

            if origen.has_style:
                destino.font = copy(origen.font)
                destino.fill = copy(origen.fill)
                destino.border = copy(origen.border)
                destino.alignment = copy(origen.alignment)
                destino.number_format = copy(origen.number_format)
                destino.protection = copy(origen.protection)

        ws.row_dimensions[fila_destino].height = \
            ws.row_dimensions[fila_origen].height


    def calcular_mediana(valores):
        valores = [v for v in valores if isinstance(v, (int, float))]
        return round(median(valores), 2) if valores else 0


    def calcular_promedio(valores):
        valores = [v for v in valores if isinstance(v, (int, float))]
        return round(sum(valores) / len(valores), 2) if valores else 0


    # =====================================================
    # CARGAR PLANTILLA
    # =====================================================

    wb = load_workbook(PLANTILLA)
    ws = wb.active
    fila_actual = 5
    if ws.max_row > 5:
        ws.delete_rows(5, ws.max_row)

    # =====================================================
    # GENERAR DATOS
    # =====================================================

    for bloque, proyectos in datos.items():

        # ---------------------------------------------
        # FILA DEL BLOQUE
        # ---------------------------------------------
        ws.cell(fila_actual, 1).value = bloque

        # Copiamos formato de la fila verde de sección
        copiar_estilo_fila(
            ws,
            fila_origen=4,
            fila_destino=fila_actual,
            max_col=52
        )

        fila_actual += 1

        fila_inicio_proyectos = fila_actual

        # ---------------------------------------------
        # PROYECTOS
        # ---------------------------------------------
        for proyecto in proyectos:

            ws.cell(fila_actual, 1).value = proyecto[0]

            for col, valor in enumerate(proyecto[1:], start=2):
                ws.cell(fila_actual, col).value = valor

            # Copia formato de una fila de proyecto
            copiar_estilo_fila(
                ws,
                fila_origen=5,
                fila_destino=fila_actual,
                max_col=52
            )

            fila_actual += 1

        fila_fin_proyectos = fila_actual - 1

        # ---------------------------------------------
        # MEDIANA
        # ---------------------------------------------
        ws.cell(fila_actual, 1).value = "MEDIANA"

        copiar_estilo_fila(
            ws,
            fila_origen=18,   
            fila_destino=fila_actual,
            max_col=52
        )

        for col in range(2, 52):

            valores = []

            for fila in range(
                fila_inicio_proyectos,
                fila_fin_proyectos + 1
            ):
                valor = ws.cell(fila, col).value

                if isinstance(valor, (int, float)):
                    valores.append(valor)

            ws.cell(fila_actual, col).value = calcular_mediana(valores)

        fila_actual += 1

        # ---------------------------------------------
        # PROMEDIO
        # ---------------------------------------------
        ws.cell(fila_actual, 1).value = "PROMEDIO"

        copiar_estilo_fila(
            ws,
            fila_origen=19,
            fila_destino=fila_actual,
            max_col=52
        )

        for col in range(2, 52):

            valores = []

            for fila in range(
                fila_inicio_proyectos,
                fila_fin_proyectos + 1
            ):
                valor = ws.cell(fila, col).value

                if isinstance(valor, (int, float)):
                    valores.append(valor)

            ws.cell(fila_actual, col).value = calcular_promedio(valores)

        fila_actual += 1

        # ---------------------------------------------
        # MAX
        # ---------------------------------------------
        ws.cell(fila_actual, 1).value = "MAX"

        copiar_estilo_fila(
            ws,
            fila_origen=20,
            fila_destino=fila_actual,
            max_col=52
        )

        for col in range(2, 52):

            valores = []

            for fila in range(
                fila_inicio_proyectos,
                fila_fin_proyectos + 1
            ):
                valor = ws.cell(fila, col).value

                if isinstance(valor, (int, float)):
                    valores.append(valor)

            ws.cell(fila_actual, col).value = (
                max(valores) if valores else 0
            )

        fila_actual += 1

        # ---------------------------------------------
        # MIN
        # ---------------------------------------------
        ws.cell(fila_actual, 1).value = "MIN"

        copiar_estilo_fila(
            ws,
            fila_origen=21,
            fila_destino=fila_actual,
            max_col=52
        )

        for col in range(2, 52):

            valores = []

            for fila in range(
                fila_inicio_proyectos,
                fila_fin_proyectos + 1
            ):
                valor = ws.cell(fila, col).value

                if isinstance(valor, (int, float)):
                    valores.append(valor)

            ws.cell(fila_actual, col).value = (
                min(valores) if valores else 0
            )

        fila_actual += 2

    # =====================================================
    # GUARDAR
    # =====================================================

    wb.save(SALIDA)

    print(f"Archivo generado: {SALIDA}")

    # ==========================================================
    # GENERAR ARCHIVO DE CORRECCIONES ACTUALIZADO
    # ==========================================================

    datos_correccion = {}

    grupos = (
        tabla_fil.index.droplevel("PROYECTO")
        .unique()
    )

    for familia_tamaño in grupos:

        bloque = tabla_fil.loc[familia_tamaño]

        proyectos = []

        for proyecto, fila in bloque.iterrows():

            proyectos.append(
                [proyecto] + fila.tolist()
            )

        datos_correccion[familia_tamaño] = proyectos

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    def generar_archivo_correcciones(
        datos,
        columnas,
        salida
    ):

        wb_corr = Workbook()
        ws_corr = wb_corr.active
        ws_corr.title = "Correcciones"

        fila = 1

        for bloque, proyectos in datos.items():

            # ----------------------------------
            # Título del bloque
            # ----------------------------------

            ws_corr.cell(fila, 1).value = bloque

            ws_corr.cell(fila, 1).font = Font(
                bold=True,
                color="FFFFFF"
            )

            ws_corr.cell(fila, 1).fill = PatternFill(
                "solid",
                fgColor="4F81BD"
            )

            fila += 1

            # ----------------------------------
            # Cabeceras
            # ----------------------------------

            ws_corr.cell(fila, 1).value = "PROYECTO"

            for c, nombre_col in enumerate(columnas, start=2):

                ws_corr.cell(
                    fila,
                    c
                ).value = nombre_col

            fila += 1

            # ----------------------------------
            # Proyectos
            # ----------------------------------

            for proyecto in proyectos:

                ws_corr.cell(
                    fila,
                    1
                ).value = proyecto[0]

                for c, valor in enumerate(
                    proyecto[1:],
                    start=2
                ):

                    ws_corr.cell(
                        fila,
                        c
                    ).value = valor

                fila += 1

            fila += 2

        wb_corr.save(salida)


    generar_archivo_correcciones(
        datos=datos_correccion,
        columnas=lista_planificaciones,
        salida="correcciones_actualizadas.xlsx"
    )

    mapeo_fase = {
        "CORTE Y ENSAMBLAJE DE CARRILES Y BANDEJAS": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "NIVELACIÓN": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "INSTALACIÓN CARRILES DE SOPORTACIÓN": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "INSTALACIÓN DE BANDEJAS": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "MECANIZADO DE ELEMENTOS DE INSTALACIÓN MECÁNICA": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "INTRODUCCIÓN/FIJACIÓN de EQUIPOS": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "ENSAMBLAJE DE CUADROS CUADRISTA": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "ENSAMBLAJE DE EQUIPOS (CUADRO PRINCIPAL Y UPS)": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "CADENAS PORTACABLES": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "INSTALACIÓN DE SUELO": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "PANELADO": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "INSTALACIÓN  DE PUERTAS": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "INSTALACIÓN DE ROXTEC": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "INSTALACIÓN ELEMENTOS y/o EQUIPOS": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "CERRAMIENTO DE ALUMINIO": "FASE DE ESTRUCTURA Y ENSAMBLAJE",
        "PERFILERIA EXTERIOR": "FASE DE ESTRUCTURA Y ENSAMBLAJE",

        "CORTE Y/O PREPARACIÓN DE CABLE": "FASE ELECTRICA",
        "MECANIZADO DE ELEMENTOS DE INSTALACIÓN ELECTRICA": "FASE ELECTRICA",
        "SISTEMA DE TIERRAS": "FASE ELECTRICA",
        "SISTEMA DE ILUMINACIÓN/EMERGENCIAS": "FASE ELECTRICA",
        "INSTALACIÓN DE CABLE": "FASE ELECTRICA",
        "CONEXIONADO DE EQUIPOS": "FASE ELECTRICA",
        "ETIQUETADO DE CABLE Y EQUIPOS": "FASE ELECTRICA",
        "TORQUEO": "FASE ELECTRICA",
        "TRANSFORMADOR": "FASE ELECTRICA",
        "TRABAJOS DE MEDIA TENSIÓN": "FASE ELECTRICA",
        "TRABAJOS DE BT EN MT": "FASE ELECTRICA",
        "PRUEBAS DE MEDIA TENSION": "FASE ELECTRICA",
        "MONTAJE DE EQUIPOS DE MONITORIZACIÓN": "FASE ELECTRICA",
        "INSTALACIÓN DE SEÑALES": "FASE ELECTRICA",
        "INSTALACIÓN DE UTP": "FASE ELECTRICA",
        "CONTROL DE ACCESO": "FASE ELECTRICA",
        "PRUEBAS ELECTRICAS SIN TENSIÓN": "FASE ELECTRICA",
        "PRUEBAS ELECTRICAS CON TENSIÓN": "FASE ELECTRICA",
        "SOPORTE A COMMISSIONING": "FASE ELECTRICA",

        "TRABAJOS DE INSTALACIÓN PCI": "FASE PCI",
        "REALIZACIÓN DE PRUEBAS PCI": "FASE PCI",

        "MONTAJE DE TUBERÍAS": "FASE CLIMATIZACION",
        "TRABAJOS DE CONEXIONADO Y VALVULERIA": "FASE CLIMATIZACION",
        "TRABAJOS DE FORRADO DE TUBERÍAS": "FASE CLIMATIZACION",
        "REALIZACIÓN DE PRUEBAS COOLING": "FASE CLIMATIZACION",

        "MODIFICACIONES FAT/FOK/WITNESS TEST ": "PUNCH LIST",

        "DESENSAMBLAJE INSTALACIÓN ELÉCTRICA": "FASE DE PREPARACION PARA ENVIO",
        "DESENSAMBLAJE INSTALACIÓN MECÁNICA": "FASE DE PREPARACION PARA ENVIO",
        "LIMPIEZA FINAL CUADROS": "FASE DE PREPARACION PARA ENVIO",
        "CIERRE DE CUADROS Y PENDIENTES ELÉCTRICOS": "FASE DE PREPARACION PARA ENVIO",
        "CIERRE DE CUADROS Y PENDIENTES MECÁNICOS": "FASE DE PREPARACION PARA ENVIO",
        "LIMPIEZA FINAL": "FASE DE PREPARACION PARA ENVIO",
        "REPASOS PINTURA": "FASE DE PREPARACION PARA ENVIO",
        "EMBALAJE": "FASE DE PREPARACION PARA ENVIO",
    }
    mapeo_grupo = {
        "CORTE Y ENSAMBLAJE DE CARRILES Y BANDEJAS": "SE M",
        "NIVELACIÓN": "Mecánico",
        "INSTALACIÓN CARRILES DE SOPORTACIÓN": "Mecánico",
        "INSTALACIÓN DE BANDEJAS": "Mecánico",
        "MECANIZADO DE ELEMENTOS DE INSTALACIÓN MECÁNICA": "Mecánico",
        "INTRODUCCIÓN/FIJACIÓN de EQUIPOS": "Mecánico",
        "ENSAMBLAJE DE CUADROS CUADRISTA": "Cuadrista",
        "ENSAMBLAJE DE EQUIPOS (CUADRO PRINCIPAL Y UPS)": "SE E",
        "CADENAS PORTACABLES": "Mecánico",
        "INSTALACIÓN DE SUELO": "Mecánico",
        "PANELADO": "Panelado",
        "INSTALACIÓN  DE PUERTAS": "Mecánico",
        "INSTALACIÓN DE ROXTEC": "Mecánico",
        "INSTALACIÓN ELEMENTOS y/o EQUIPOS": "Mecánico",
        "CERRAMIENTO DE ALUMINIO": "Cerramiento",
        "PERFILERIA EXTERIOR": "Perfileria",
        "CORTE Y/O PREPARACIÓN DE CABLE": "SE M",
        "MECANIZADO DE ELEMENTOS DE INSTALACIÓN ELECTRICA": "Eléctrico",
        "SISTEMA DE TIERRAS": "Eléctrico",
        "SISTEMA DE ILUMINACIÓN/EMERGENCIAS": "Eléctrico",
        "INSTALACIÓN DE CABLE": "Eléctrico",
        "CONEXIONADO DE EQUIPOS": "Eléctrico",
        "ETIQUETADO DE CABLE Y EQUIPOS": "Eléctrico",
        "TORQUEO": "SE E",
        "TRANSFORMADOR": "Eléctrico",
        "TRABAJOS DE MEDIA TENSIÓN": "Eléctrico",
        "TRABAJOS DE BT EN MT": "Eléctrico",
        "PRUEBAS DE MEDIA TENSION": "Eléctrico",
        "MONTAJE DE EQUIPOS DE MONITORIZACIÓN": "Eléctrico",
        "INSTALACIÓN DE SEÑALES": "Eléctrico",
        "INSTALACIÓN DE UTP": "Eléctrico",
        "CONTROL DE ACCESO": "Eléctrico",
        "PRUEBAS ELECTRICAS SIN TENSIÓN": "Eléctrico",
        "PRUEBAS ELECTRICAS CON TENSIÓN": "Eléctrico",
        "SOPORTE A COMMISSIONING": "Eléctrico",
        "TRABAJOS DE INSTALACIÓN PCI": "PCI",
        "REALIZACIÓN DE PRUEBAS PCI": "PCI",
        "MONTAJE DE TUBERÍAS": "Climatización",
        "TRABAJOS DE CONEXIONADO Y VALVULERIA": "Climatización",
        "TRABAJOS DE FORRADO DE TUBERÍAS": "Climatización",
        "REALIZACIÓN DE PRUEBAS COOLING": "Climatización",
        "DESENSAMBLAJE INSTALACIÓN ELÉCTRICA": "Eléctrico",
        "DESENSAMBLAJE INSTALACIÓN MECÁNICA": "Mecánico",
        "LIMPIEZA FINAL CUADROS": "Limpieza",
        "CIERRE DE CUADROS Y PENDIENTES ELÉCTRICOS": "Eléctrico",
        "CIERRE DE CUADROS Y PENDIENTES MECÁNICOS": "Mecánico",
        "LIMPIEZA FINAL": "Limpieza",
        "REPASOS PINTURA": "SE M",
        "EMBALAJE": "Embalaje"
    }
    grupos_resumen = [
        "Mecánico",
        "Eléctrico",
        "Cuadrista",
        "SE M",
        "SE E",
        "Cerramiento",
        "Perfileria",
        "Limpieza",
        "Embalaje",
        "PCI",
        "Panelado",
        "Climatización"
    ]

    # =====================================================
    # FAMILIAS
    # =====================================================

    familias = list(
        tabla.index
        .get_level_values("FAMILIA_TAMAÑO")
        .unique()
    )

    # =====================================================
    # PROMEDIOS POR TAREA Y FAMILIA
    # =====================================================

    promedios_tareas = {}

    for tarea in mapeo_grupo:

        if tarea not in tabla.columns:
            continue

        promedios_tareas[tarea] = {}

        for familia in familias:

            bloque = tabla.xs(
                familia,
                level="FAMILIA_TAMAÑO"
            )

            promedio = round(
                bloque[tarea].mean(),
                2
            )

            promedios_tareas[tarea][familia] = promedio

    # =====================================================
    # RESUMEN POR GRUPOS
    # =====================================================

    resumen = {}

    for grupo in grupos_resumen:

        resumen[grupo] = {}

        for familia in familias:
            resumen[grupo][familia] = 0

    for tarea, valores in promedios_tareas.items():

        grupo = mapeo_grupo.get(tarea)

        if grupo is None:
            continue

        for familia in familias:

            resumen[grupo][familia] += (
                valores.get(familia, 0)
            )

    # =====================================================
    # CREAR HOJA RESUMEN
    # =====================================================

    if "Resumen" in wb.sheetnames:
        del wb["Resumen"]

    ws_resumen = wb.create_sheet(
        "Resumen",
        0
    )

    # =====================================================
    # TABLA SUPERIOR
    # =====================================================

    ws_resumen.cell(1, 1).value = "GRUPO"

    for col, familia in enumerate(
        familias,
        start=2
    ):

        ws_resumen.cell(
            1,
            col
        ).value = familia

    fila = 2

    for grupo in grupos_resumen:

        ws_resumen.cell(
            fila,
            1
        ).value = grupo

        for col, familia in enumerate(
            familias,
            start=2
        ):

            ws_resumen.cell(
                fila,
                col
            ).value = round(
                resumen[grupo][familia],
                2
            )

        fila += 1

    # =====================================================
    # TOTAL
    # =====================================================

    ws_resumen.cell(
        fila,
        1
    ).value = "TOTAL"

    for col, familia in enumerate(
        familias,
        start=2
    ):

        total = sum(
            resumen[g][familia]
            for g in grupos_resumen
        )

        ws_resumen.cell(
            fila,
            col
        ).value = round(
            total,
            2
        )

    fila_total = fila

    # =====================================================
    # TABLA INFERIOR
    # =====================================================

    fila_detalle = fila_total + 4

    ws_resumen.cell(
        fila_detalle,
        1
    ).value = "GRUPO"

    ws_resumen.cell(
        fila_detalle,
        2
    ).value = "FASE"

    ws_resumen.cell(
        fila_detalle,
        3
    ).value = "TAREA"

    for col, familia in enumerate(
        familias,
        start=4
    ):

        ws_resumen.cell(
            fila_detalle,
            col
        ).value = familia

    fila_detalle += 1

    for tarea in mapeo_grupo:

        if tarea not in promedios_tareas:
            continue

        ws_resumen.cell(
            fila_detalle,
            1
        ).value = mapeo_grupo[tarea]

        ws_resumen.cell(
            fila_detalle,
            2
        ).value = mapeo_fase.get(
            tarea,
            ""
        )

        ws_resumen.cell(
            fila_detalle,
            3
        ).value = tarea

        for col, familia in enumerate(
            familias,
            start=4
        ):

            ws_resumen.cell(
                fila_detalle,
                col
            ).value = promedios_tareas[tarea].get(
                familia,
                0
            )

        fila_detalle += 1

    # =====================================================
    # FORMATO
    # =====================================================

    from openpyxl.styles import Font

    negrita = Font(
        bold=True
    )

    for celda in ws_resumen[1]:
        celda.font = negrita

    for celda in ws_resumen[fila_total]:
        celda.font = negrita

    for celda in ws_resumen[fila_total + 4]:
        celda.font = negrita

    # =====================================================
    # AJUSTAR COLUMNAS
    # =====================================================

    for columna in ws_resumen.columns:

        ancho = max(
            len(str(c.value))
            if c.value is not None else 0
            for c in columna
        )

        ws_resumen.column_dimensions[
            columna[0].column_letter
        ].width = ancho + 4

    # =====================================================
    # GUARDAR
    # =====================================================

    wb.save(SALIDA)

    print(
        f"Archivo generado correctamente: {SALIDA}"
    )

    return (
        "trabajo_patron_horas.xlsx",
        "correcciones_actualizadas.xlsx"
    )
